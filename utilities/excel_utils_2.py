import pytest
from openpyxl import load_workbook

class ExcelUtils:
    @staticmethod
    def load_excel_data(file_path, sheet_name):
        try:
            # Open the workbook and select the sheet
            workbook = load_workbook(file_path)
            sheet = workbook[sheet_name]

            # Extract data from the sheet
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                data.append(row)
            return data
        except FileNotFoundError:
            raise FileNotFoundError(f"The file {file_path} was not found.")
        except KeyError:
            raise KeyError(f"The sheet {sheet_name} does not exist in the workbook.")
